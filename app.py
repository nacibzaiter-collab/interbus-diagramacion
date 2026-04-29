import os, io, re, json
from datetime import datetime
from flask import Flask, request, render_template, jsonify, send_file, session
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import copy

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "interbus-2025-secret")

# ── Carpeta de uploads temporal ──────────────────────────────────────────────
UPLOAD_FOLDER = "/tmp/interbus_uploads"
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# ── Tabla de legajos ─────────────────────────────────────────────────────────
LEGAJOS_DEFAULT = {
    "2040": "AGUERRE P.",
    "2044": "BELTRAMO J.",
    "2037": "DEL CANTO M.",
    "2035": "DEL CANTO M.",
    "2050": "FERREYRA G.",
    "2011": "GIULIANI M.",
    "2056": "KANCHEFF C.",
    "2041": "LAMBRUSCHI J.",
    "2017": "LASO M.",
    "2057": "LUQUEZ D.",
    "2038": "MURCIA S.",
    "2022": "PAOLLICELLI A.",
    "2048": "PASCASIO G.",
    "2015": "PEREYRA W.",
    "2003": "ROSCHINI M.",
    "2452": "RUIZ H.",
    "2460": "VESPRINI R.",
    "2046": "VOLPINI A.",
    "3003": "PACI G.",
    "2111": "GUARDIA",
}

MESES = ["ene","feb","mar","abr","may","jun","jul","ago","sep","oct","nov","dic"]

def get_apellido(codigo, legajos):
    return legajos.get(str(codigo).strip(), str(codigo))

def es_legajo_valido(val):
    s = str(val or "").strip()
    return bool(re.match(r'^\d{4,5}$', s)) and int(s) > 1000

def es_guardia(val):
    s = str(val or "").strip().upper()
    return s in ("", "****", "GUARDIA", "ALVARELLOS")

def fmt_fecha(fecha_str):
    """2026-04-25 -> 25-abr"""
    m = re.match(r'^(\d{4})-(\d{2})-(\d{2})', str(fecha_str))
    if m:
        d, mo = int(m.group(3)), int(m.group(2))-1
        return f"{d:02d}-{MESES[mo]}"
    return str(fecha_str)

def fmt_fecha_larga(fecha_str):
    """2026-04-25 -> 25 ABR 2026"""
    m = re.match(r'^(\d{4})-(\d{2})-(\d{2})', str(fecha_str))
    if m:
        d, mo, y = int(m.group(3)), int(m.group(2))-1, m.group(1)
        return f"{d:02d} {MESES[mo].upper()} {y}"
    return str(fecha_str).upper()

# ────────────────────────────────────────────────────────────────────────────
# PARSEAR PLANILLA DE CITACIÓN
# Col A(0)=svc ALV, B(1)=cond ALV, C(2)=ayud ALV
# Col E(4)=FECHA
# Col G(6)=svc ACE, H(7)=cond ACE, I(8)=ayud ACE
# ────────────────────────────────────────────────────────────────────────────
def parsear_citacion(wb):
    result = {}  # { "2026-04-25": [{ruta,svc,legajo,tipo}] }

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        fecha_actual = None

        for row in ws.iter_rows(values_only=True):
            row = [str(c) if c is not None else "" for c in row]
            # Extender si la fila es corta
            while len(row) < 10:
                row.append("")

            # Detectar fecha en col E (índice 4)
            fc = row[4].strip()
            if fc:
                m = re.match(r'^(\d{4}-\d{2}-\d{2})', fc)
                if m:
                    fecha_actual = m.group(1)
                    if fecha_actual not in result:
                        result[fecha_actual] = []
                    continue
                # Formato "25/04/2026" o "25-04-2026"
                m2 = re.match(r'^(\d{2})[/\-](\d{2})[/\-](\d{4})', fc)
                if m2:
                    fecha_actual = f"{m2.group(3)}-{m2.group(2)}-{m2.group(1)}"
                    if fecha_actual not in result:
                        result[fecha_actual] = []
                    continue

            if not fecha_actual:
                continue

            svc_a = row[0].strip()
            cond_a = row[1].strip()
            ayud_a = row[2].strip()
            svc_c = row[6].strip()
            cond_c = row[7].strip()
            ayud_c = row[8].strip()

            # ALVAREZ: conductor=PRIMERA, ayudante=RELEVO
            if svc_a and re.match(r'^\d+$', svc_a):
                if es_legajo_valido(cond_a):
                    result[fecha_actual].append({
                        "ruta": "ALVAREZ", "svc": int(svc_a),
                        "legajo": cond_a, "tipo": "PRIMERA"
                    })
                if es_legajo_valido(ayud_a):
                    result[fecha_actual].append({
                        "ruta": "ALVAREZ", "svc": int(svc_a),
                        "legajo": ayud_a, "tipo": "RELEVO"
                    })

            # ACEBAL: conductor=PRIMERA, ayudante=RELEVO
            if svc_c and re.match(r'^\d+$', svc_c):
                if not es_guardia(cond_c) and es_legajo_valido(cond_c):
                    result[fecha_actual].append({
                        "ruta": "ACEBAL", "svc": int(svc_c),
                        "legajo": cond_c, "tipo": "PRIMERA"
                    })
                if es_legajo_valido(ayud_c):
                    result[fecha_actual].append({
                        "ruta": "ACEBAL", "svc": int(svc_c),
                        "legajo": ayud_c, "tipo": "RELEVO"
                    })

    return result

# ────────────────────────────────────────────────────────────────────────────
# PARSEAR MAESTRO: extraer estructura de slots
# ────────────────────────────────────────────────────────────────────────────
def parsear_maestro(wb):
    """
    Devuelve dict: { sheet_name: { dias_count, slots: { "diaOrd,RUTA,svc,TIPO": {coche,dia,mes,anio,...} } } }
    """
    maps = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append([str(c) if c is not None else "" for c in row])

        # Encontrar todas las filas con "PLANILLA" en col B (índice 1)
        all_starts = [i for i, r in enumerate(rows) if len(r) > 1 and r[1] == "PLANILLA"]
        bases = [s - 4 for s in all_starts]

        # Mapear días únicos
        dia_map = {}
        for base in bases:
            if base + 7 >= len(rows):
                continue
            f7 = rows[base + 7]
            key = f"{f7[1] if len(f7) > 1 else ''}/{f7[2] if len(f7) > 2 else ''}"
            if key not in dia_map:
                dia_map[key] = len(dia_map)

        slots = {}
        for base in bases:
            if base + 9 >= len(rows):
                continue
            f0 = rows[base]
            f4 = rows[base + 4]
            f7 = rows[base + 7]
            f9 = rows[base + 9]

            desc = (f0[15] if len(f0) > 15 else "").upper()
            dia_key = f"{f7[1] if len(f7) > 1 else ''}/{f7[2] if len(f7) > 2 else ''}"
            dia_ord = dia_map.get(dia_key, 0)

            ruta = "ALVAREZ" if "LVAR" in desc else ("ACEBAL" if "ACEBAL" in desc else "")
            tipo = "NOCTURNO" if "NOCTURNO" in desc else ("RELEVO" if "RELEVO" in desc else "PRIMERA")
            sm = re.search(r'S(\d+)', desc)
            snum = int(sm.group(1)) if sm else 1

            if ruta:
                k = f"{dia_ord},{ruta},{snum},{tipo}"
                slots[k] = {
                    "coche":     f4[9]  if len(f4) > 9  else "",
                    "dia":       f7[1]  if len(f7) > 1  else "",
                    "mes":       f7[2]  if len(f7) > 2  else "",
                    "anio":      f7[3]  if len(f7) > 3  else "",
                    "conductor": f9[1]  if len(f9) > 1  else "",
                    "legajo":    f9[3]  if len(f9) > 3  else "",
                    "base":      base,
                    "desc":      f0[15] if len(f0) > 15 else "",
                }

        maps[sheet_name] = {
            "dias_count": len(dia_map),
            "slots": slots,
        }

    return maps

# ────────────────────────────────────────────────────────────────────────────
# GENERAR PLANILLAS: combinar citación + maestro
# ────────────────────────────────────────────────────────────────────────────
def generar_planillas(cit_data, maestro_maps, sheet_name, contador_inicio, legajos, coches):
    mapa = maestro_maps.get(sheet_name)
    if not mapa:
        return [], f"Pestaña '{sheet_name}' no encontrada"

    slots = mapa["slots"]
    dias_count = mapa["dias_count"]
    fechas = sorted(cit_data.keys())
    planillas = []
    num = contador_inicio
    log = []

    for fi, fecha in enumerate(fechas):
        dia_ord = fi % dias_count
        svcs = cit_data[fecha]
        coches_usados = []
        log.append(f"\n📅 {fmt_fecha_larga(fecha)} (día maestro {dia_ord})")

        for svc in svcs:
            apellido = get_apellido(svc["legajo"], legajos)
            k = f"{dia_ord},{svc['ruta']},{svc['svc']},{svc['tipo']}"
            slot = slots.get(k)

            if slot:
                coche = slot["coche"]
            else:
                libres = [c for c in coches if c not in coches_usados]
                coche = libres[0] if libres else (coches[0] if coches else "??")
                log.append(f"  ⚠ Sin slot maestro: {k}")

            coches_usados.append(coche)

            planillas.append({
                "nro":       num,
                "fecha":     fecha,
                "fecha_disp": fmt_fecha(fecha),
                "legajo":    svc["legajo"],
                "conductor": apellido,
                "ruta":      "ROSARIO - " + svc["ruta"],
                "svc":       svc["svc"],
                "tipo":      svc["tipo"],
                "coche":     coche,
                "pestaña":   sheet_name,
            })
            log.append(f"  ✓ #{num}  {apellido:<18} | {svc['ruta']} S{svc['svc']} {svc['tipo']} | Coche {coche}")
            num += 1

    log.append(f"\n✅ {len(planillas)} planillas. Próxima: #{num}")
    return planillas, num, "\n".join(log)

# ────────────────────────────────────────────────────────────────────────────
# GENERAR EXCEL DE RESULTADO
# ────────────────────────────────────────────────────────────────────────────
def generar_excel(planillas):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen"

    # Estilos
    hdr_font  = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    hdr_fill  = PatternFill("solid", fgColor="1A2539")
    gold_font = Font(bold=True, color="C8A020", name="Calibri", size=11)
    body_font = Font(name="Calibri", size=11)
    center    = Alignment(horizontal="center", vertical="center")
    thin      = Side(style="thin", color="2D3A52")
    border    = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ["N° Planilla", "Fecha", "Legajo", "Conductor", "Ruta", "Servicio", "Tipo", "Coche", "Pestaña"]
    col_w   = [12, 10, 9, 20, 24, 9, 10, 9, 26]

    for col, (h, w) in enumerate(zip(headers, col_w), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = center
        cell.border = border
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.row_dimensions[1].height = 20

    alt_fill = PatternFill("solid", fgColor="1A2539")
    for i, p in enumerate(planillas, 2):
        vals = [p["nro"], p["fecha_disp"], p["legajo"], p["conductor"],
                p["ruta"], p["svc"], p["tipo"], f"Coche {p['coche']}", p["pestaña"]]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.font = gold_font if col == 1 else body_font
            cell.border = border
            cell.alignment = center
            if i % 2 == 0:
                cell.fill = alt_fill

    # Hojas por fecha
    por_fecha = {}
    for p in planillas:
        por_fecha.setdefault(p["fecha"], []).append(p)

    for fecha, items in sorted(por_fecha.items()):
        sn = fmt_fecha(fecha).replace("/", "-")[:31]
        ws2 = wb.create_sheet(title=sn)
        for col, (h, w) in enumerate(zip(headers, col_w), 1):
            cell = ws2.cell(row=1, column=col, value=h)
            cell.font = hdr_font; cell.fill = hdr_fill
            cell.alignment = center; cell.border = border
            ws2.column_dimensions[get_column_letter(col)].width = w
        for i, p in enumerate(items, 2):
            vals = [p["nro"], p["fecha_disp"], p["legajo"], p["conductor"],
                    p["ruta"], p["svc"], p["tipo"], f"Coche {p['coche']}", p["pestaña"]]
            for col, val in enumerate(vals, 1):
                cell = ws2.cell(row=i, column=col, value=val)
                cell.font = gold_font if col == 1 else body_font
                cell.border = border; cell.alignment = center
                if i % 2 == 0:
                    cell.fill = alt_fill

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

# ════════════════════════════════════════════════════════════════════════════
# RUTAS FLASK
# ════════════════════════════════════════════════════════════════════════════

MAESTRO_PATH = os.path.join(UPLOAD_FOLDER, "maestro_actual.xlsm")

@app.route("/")
def index():
    maestro_ok = os.path.exists(MAESTRO_PATH)
    maestro_info = ""
    if maestro_ok:
        ts = os.path.getmtime(MAESTRO_PATH)
        dt = datetime.fromtimestamp(ts)
        maestro_info = f"Planillas Maestro {dt.month:02d}/{str(dt.year)[-2:]}"
    return render_template("index.html",
                           maestro_ok=maestro_ok,
                           maestro_info=maestro_info)

@app.route("/upload_maestro", methods=["POST"])
def upload_maestro():
    f = request.files.get("maestro")
    if not f:
        return jsonify({"ok": False, "error": "Sin archivo"}), 400
    f.save(MAESTRO_PATH)
    now = datetime.now()
    nombre = f"Planillas Maestro {now.month:02d}/{str(now.year)[-2:]}"
    return jsonify({"ok": True, "nombre": nombre})

@app.route("/get_sheets")
def get_sheets():
    if not os.path.exists(MAESTRO_PATH):
        return jsonify({"ok": False, "error": "No hay Maestro cargado"}), 400
    try:
        wb = load_workbook(MAESTRO_PATH, read_only=True, keep_vba=False)
        return jsonify({"ok": True, "sheets": wb.sheetnames})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

@app.route("/generar", methods=["POST"])
def generar():
    # Validar entradas
    cit_file   = request.files.get("citacion")
    sheet_name = request.form.get("sheet_name", "").strip()
    contador   = int(request.form.get("contador", 6800))
    legajos_json = request.form.get("legajos", "{}")
    coches_json  = request.form.get("coches", "[]")

    if not cit_file:
        return jsonify({"ok": False, "error": "Falta la planilla de citación"}), 400
    if not os.path.exists(MAESTRO_PATH):
        return jsonify({"ok": False, "error": "No hay Maestro cargado. Subilo primero."}), 400
    if not sheet_name:
        return jsonify({"ok": False, "error": "Seleccioná el tipo de semana"}), 400

    try:
        legajos = json.loads(legajos_json)
        coches  = json.loads(coches_json)
    except:
        legajos = LEGAJOS_DEFAULT
        coches  = ["02","03","04","05","06","07","09","10","11"]

    try:
        # Leer archivos
        cit_wb  = load_workbook(io.BytesIO(cit_file.read()), data_only=True)
        mae_wb  = load_workbook(MAESTRO_PATH, keep_vba=False)

        cit_data     = parsear_citacion(cit_wb)
        maestro_maps = parsear_maestro(mae_wb)

        if not cit_data:
            return jsonify({"ok": False, "error": "No se detectaron fechas en la citación. Verificá el formato del archivo."}), 400

        planillas, nuevo_contador, log_txt = generar_planillas(
            cit_data, maestro_maps, sheet_name, contador, legajos, coches
        )

        # Guardar en sesión para download
        session["planillas"]      = planillas
        session["nuevo_contador"] = nuevo_contador

        # Armar resumen por fecha para el frontend
        por_fecha = {}
        for p in planillas:
            por_fecha.setdefault(p["fecha"], []).append(p)

        fechas_resumen = [
            {
                "fecha":     k,
                "fecha_disp": fmt_fecha_larga(k),
                "cantidad":  len(v),
                "items":     v,
            }
            for k, v in sorted(por_fecha.items())
        ]

        return jsonify({
            "ok":             True,
            "total":          len(planillas),
            "primera":        planillas[0]["nro"]  if planillas else 0,
            "ultima":         planillas[-1]["nro"] if planillas else 0,
            "nuevo_contador": nuevo_contador,
            "log":            log_txt,
            "fechas":         fechas_resumen,
        })

    except Exception as e:
        import traceback
        return jsonify({"ok": False, "error": str(e), "trace": traceback.format_exc()}), 500

@app.route("/descargar_excel")
def descargar_excel():
    planillas = session.get("planillas", [])
    if not planillas:
        return "Sin datos", 400
    buf = generar_excel(planillas)
    nombre = f"Planillas_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(buf, as_attachment=True, download_name=nombre,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.route("/descargar_csv")
def descargar_csv():
    planillas = session.get("planillas", [])
    if not planillas:
        return "Sin datos", 400
    lines = ["N°,Fecha,Legajo,Conductor,Ruta,Servicio,Tipo,Coche"]
    for p in planillas:
        lines.append(f"{p['nro']},{p['fecha_disp']},{p['legajo']},\"{p['conductor']}\","
                     f"\"{p['ruta']}\",{p['svc']},{p['tipo']},{p['coche']}")
    buf = io.BytesIO("\n".join(lines).encode("utf-8"))
    nombre = f"Planillas_{datetime.now().strftime('%Y%m%d_%H%M')}.csv"
    return send_file(buf, as_attachment=True, download_name=nombre, mimetype="text/csv")

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=False)
